#!/usr/bin/env python3
"""
Descarga y procesa fuentes externas CONEVAL y SESNSP para el panel ML.

Uso:
  python3 fetch_external_data.py
  python3 fetch_external_data.py --solo-sesnsp
  python3 fetch_external_data.py --solo-coneval

Salidas:
  data/external/coneval_estatal.csv
  data/external/sesnsp_victimas_estatal.csv

Tras ejecutar, correr: python3 build_state_panel.py
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import unicodedata
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent
EXTERNAL = ROOT / "data" / "external"
RAW = EXTERNAL / "raw"
CONEVAL_OUT = EXTERNAL / "coneval_estatal.csv"
SESNSP_OUT = EXTERNAL / "sesnsp_victimas_estatal.csv"

CONEVAL_ZIPS = {
    "2008_2018": (
        "https://www.coneval.org.mx/Medicion/MP/Documents/Pobreza_18/AE_estatal_2008_2018.zip",
        RAW / "coneval" / "AE_estatal_2008_2018.zip",
    ),
    "2016_2020": (
        "https://www.coneval.org.mx/Medicion/MP/Documents/MMP_2018_2020/AE_estatal_2016_2020.zip",
        RAW / "coneval" / "AE_estatal_2016_2020.zip",
    ),
    "2022": (
        "https://www.coneval.org.mx/Medicion/MP/Documents/MMP_2022/AE_estatal_2022.zip",
        RAW / "coneval" / "AE_estatal_2022.zip",
    ),
}

SESNSP_VICTIMAS_URL = (
    "https://raw.githubusercontent.com/lapanquecita/incidencia-delictiva/main/data/victimas.csv"
)
SESNSP_VICTIMAS_N_URL = (
    "https://raw.githubusercontent.com/lapanquecita/incidencia-delictiva/main/data/victimas_n.csv"
)

STATE_NAMES: dict[str, str] = {
    "01": "AGUASCALIENTES",
    "02": "BAJA CALIFORNIA",
    "03": "BAJA CALIFORNIA SUR",
    "04": "CAMPECHE",
    "05": "COAHUILA",
    "06": "COLIMA",
    "07": "CHIAPAS",
    "08": "CHIHUAHUA",
    "09": "CIUDAD DE MEXICO",
    "10": "DURANGO",
    "11": "GUANAJUATO",
    "12": "GUERRERO",
    "13": "HIDALGO",
    "14": "JALISCO",
    "15": "ESTADO DE MEXICO",
    "16": "MICHOACAN",
    "17": "MORELOS",
    "18": "NAYARIT",
    "19": "NUEVO LEON",
    "20": "OAXACA",
    "21": "PUEBLA",
    "22": "QUERETARO",
    "23": "QUINTANA ROO",
    "24": "SAN LUIS POTOSI",
    "25": "SINALOA",
    "26": "SONORA",
    "27": "TABASCO",
    "28": "TAMAULIPAS",
    "29": "TLAXCALA",
    "30": "VERACRUZ",
    "31": "YUCATAN",
    "32": "ZACATECAS",
}

SHEET_TO_CVE: dict[str, str] = {
    "aguascalientes": "01",
    "baja california": "02",
    "baja california sur": "03",
    "campeche": "04",
    "coahuila": "05",
    "coahuila de zaragoza": "05",
    "colima": "06",
    "chiapas": "07",
    "chihuahua": "08",
    "ciudad de mexico": "09",
    "cuidad de mexico": "09",  # typo en anexo CONEVAL
    "durango": "10",
    "guanajuato": "11",
    "guerrero": "12",
    "hidalgo": "13",
    "jalisco": "14",
    "estado de mexico": "15",
    "mexico": "15",
    "méxico": "15",
    "michoacan": "16",
    "michoacán": "16",
    "michoacan de ocampo": "16",
    "michoacán de ocampo": "16",
    "morelos": "17",
    "nayarit": "18",
    "nuevo leon": "19",
    "nuevo león": "19",
    "oaxaca": "20",
    "puebla": "21",
    "queretaro": "22",
    "querétaro": "22",
    "quintana roo": "23",
    "san luis potosi": "24",
    "san luis potosí": "24",
    "sinaloa": "25",
    "sonora": "26",
    "tabasco": "27",
    "tamaulipas": "28",
    "tlaxcala": "29",
    "veracruz": "30",
    "veracruz de ignacio de la llave": "30",
    "yucatan": "31",
    "yucatán": "31",
    "zacatecas": "32",
}

MESES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

CRIMES = {
    "tasa_hom_doloso": "Homicidio doloso",
    "tasa_secuestro": "Secuestro",
    "tasa_feminicidio": "Feminicidio",
    "tasa_extorsion": "Extorsión",
}


def norm_key(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", text.lower().strip())


def download(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 1000:
        return
    print(f"Descargando {dest.name}...")
    urllib.request.urlretrieve(url, dest)


def parse_year(val: Any) -> int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        y = int(val)
        return y if 1990 <= y <= 2035 else None
    s = str(val).strip().replace("*", "")
    if s.isdigit():
        y = int(s)
        return y if 1990 <= y <= 2035 else None
    return None


def find_indicator_rows(rows: list[tuple[Any, ...]]) -> dict[str, dict[int, float]]:
    """Extrae pobreza, pobreza extrema y carencias por año desde hoja estatal."""
    year_cols: dict[int, int] = {}
    out: dict[str, dict[int, float]] = {
        "pobreza_pct": {},
        "pobreza_ext_pct": {},
        "carencias_pct": {},
    }

    pct_start = None
    pct_end = None
    for row in rows[:15]:
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            txt = norm_key(str(cell))
            if txt == "porcentaje":
                pct_start = col_idx
            if "miles de personas" in txt:
                pct_end = col_idx
                break
        if pct_start is not None and pct_end is not None:
            break

    # localizar fila de años en bloque de porcentajes
    for row in rows[:20]:
        cols: dict[int, int] = {}
        for col_idx, cell in enumerate(row):
            if pct_start is not None and col_idx < pct_start:
                continue
            if pct_end is not None and col_idx >= pct_end:
                continue
            y = parse_year(cell)
            if y is not None:
                cols[y] = col_idx
        if len(cols) >= 2:
            year_cols = cols
            break

    if not year_cols:
        return out

    targets_ordered = [
        ("pobreza_ext_pct", "poblacion en situacion de pobreza extrema"),
        ("pobreza_pct", "poblacion en situacion de pobreza"),
        ("carencias_pct", "poblacion con al menos una carencia social"),
    ]

    for row in rows:
        labels = [norm_key(str(c)) for c in row[:4] if c is not None and str(c).strip()]
        if not labels:
            continue
        label_joined = " ".join(labels)
        for var, needle in targets_ordered:
            if needle not in label_joined:
                continue
            if var == "pobreza_pct" and "extrema" in label_joined:
                continue
            if var == "pobreza_pct" and "moderada" in label_joined:
                continue
            for year, col in year_cols.items():
                val = row[col] if col < len(row) else None
                if isinstance(val, (int, float)) and not (
                    isinstance(val, float) and math.isnan(val)
                ):
                    out[var][year] = float(val)

    return out


def parse_coneval_xlsx(xlsx_path: Path) -> dict[str, dict[str, dict[int, float]]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result: dict[str, dict[str, dict[int, float]]] = {}
    for sheet in wb.sheetnames:
        if sheet in ("Contenido", "Estados Unidos Mexicanos"):
            continue
        cve = SHEET_TO_CVE.get(norm_key(sheet))
        if not cve:
            print(f"  Advertencia: hoja sin mapeo CVE: {sheet}")
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        parsed = find_indicator_rows(rows)
        result[cve] = parsed
    wb.close()
    return result


def merge_coneval_panels(
    panels: list[dict[str, dict[str, dict[int, float]]]]
) -> dict[str, dict[int, dict[str, float]]]:
    merged: dict[str, dict[int, dict[str, float]]] = defaultdict(
        lambda: defaultdict(dict)
    )
    for panel in panels:
        for cve, vars_by_year in panel.items():
            for var, year_vals in vars_by_year.items():
                for year, val in year_vals.items():
                    merged[cve][year][var] = val
    return merged


def interpolate_years(
    sparse: dict[str, dict[int, dict[str, float]]],
    desde: int,
    hasta: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for cve in sorted(STATE_NAMES):
        by_year = sparse.get(cve, {})
        measured_years = sorted(by_year.keys())
        for anio in range(desde, hasta + 1):
            rec: dict[str, Any] = {
                "cve_estado": cve,
                "estado": STATE_NAMES[cve],
                "anio": anio,
                "fuente": "CONEVAL anexos estatales",
            }
            for var in ("pobreza_pct", "pobreza_ext_pct", "carencias_pct", "gini"):
                if anio in by_year and var in by_year[anio]:
                    rec[var] = round(by_year[anio][var], 4)
                elif measured_years:
                    lower = max((y for y in measured_years if y <= anio), default=None)
                    upper = min((y for y in measured_years if y >= anio), default=None)
                    v = None
                    if lower is not None and upper is not None:
                        if lower == upper:
                            v = by_year[lower].get(var)
                        else:
                            v0 = by_year[lower].get(var)
                            v1 = by_year[upper].get(var)
                            if v0 is not None and v1 is not None:
                                frac = (anio - lower) / (upper - lower)
                                v = v0 + frac * (v1 - v0)
                    elif lower is not None:
                        # carry-forward del último año medido
                        v = by_year[lower].get(var)
                    elif upper is not None:
                        v = by_year[upper].get(var)
                    if v is not None:
                        rec[var] = round(v, 4)
                    else:
                        rec[var] = ""
                else:
                    rec[var] = ""
            rows.append(rec)
    return rows


def process_coneval(desde: int, hasta: int) -> None:
    extract_dir = RAW / "coneval" / "extracted"
    extract_dir.mkdir(parents=True, exist_ok=True)
    panels = []
    for _, (url, dest) in CONEVAL_ZIPS.items():
        download(url, dest)
        with zipfile.ZipFile(dest) as zf:
            xlsx_name = zf.namelist()[0]
            xlsx_path = extract_dir / Path(xlsx_name).name
            if not xlsx_path.exists():
                zf.extract(xlsx_name, extract_dir)
                extracted = extract_dir / xlsx_name
                if extracted != xlsx_path:
                    extracted.rename(xlsx_path)
        print(f"Procesando {xlsx_path.name}...")
        panels.append(parse_coneval_xlsx(xlsx_path))

    merged = merge_coneval_panels(panels)
    rows = interpolate_years(merged, desde, hasta)

    with CONEVAL_OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "cve_estado",
                "estado",
                "anio",
                "pobreza_pct",
                "pobreza_ext_pct",
                "carencias_pct",
                "gini",
                "fuente",
            ],
        )
        w.writeheader()
        w.writerows(rows)
    print(f"CONEVAL: {len(rows)} filas → {CONEVAL_OUT}")


def load_poblacion() -> dict[tuple[str, int], int]:
    pob_path = EXTERNAL / "poblacion_estatal.csv"
    out: dict[tuple[str, int], int] = {}
    if not pob_path.exists():
        return out
    with pob_path.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[(row["cve_estado"].zfill(2), int(row["anio"]))] = int(row["poblacion"])
    return out


def parse_int(val: Any) -> int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    return int(str(val).replace(",", "").strip() or 0)


def process_sesnsp(desde: int, hasta: int) -> None:
    victimas_path = RAW / "sesnsp_victimas.csv"
    victimas_n_path = RAW / "sesnsp_victimas_n.csv"
    download(SESNSP_VICTIMAS_URL, victimas_path)
    # victimas_n opcional (metodología 2026+)
    try:
        download(SESNSP_VICTIMAS_N_URL, victimas_n_path)
    except Exception:
        victimas_n_path = None

    counts: dict[tuple[str, int, str], int] = defaultdict(int)
    sources = [victimas_path]
    if victimas_n_path and victimas_n_path.exists():
        sources.append(victimas_n_path)

    for path in sources:
        with path.open(encoding="latin-1") as f:
            reader = csv.DictReader(f)
            for row in reader:
                anio = int(row["Año"])
                if anio < desde or anio > hasta:
                    continue
                subtipo = row["Subtipo de delito"]
                if subtipo not in CRIMES.values():
                    continue
                cve = str(int(row["Clave_Ent"])).zfill(2)
                total = sum(parse_int(row.get(m)) for m in MESES)
                counts[(cve, anio, subtipo)] += total

    poblacion = load_poblacion()
    if not poblacion:
        # importar interpolación desde build_state_panel
        from build_state_panel import interpolate_poblacion

        for cve in STATE_NAMES:
            for anio in range(desde, hasta + 1):
                poblacion[(cve, anio)] = interpolate_poblacion(cve, anio)

    rows: list[dict[str, Any]] = []
    for cve in sorted(STATE_NAMES):
        for anio in range(desde, hasta + 1):
            pob = poblacion.get((cve, anio), 0) or 1
            rec: dict[str, Any] = {
                "cve_estado": cve,
                "estado": STATE_NAMES[cve],
                "anio": anio,
                "fuente": "SESNSP víctimas fuero común (datos.gob.mx vía incidencia-delictiva)",
            }
            for var, subtipo in CRIMES.items():
                n = counts.get((cve, anio, subtipo), 0)
                rec[var] = round(100_000 * n / pob, 4) if n else 0.0
            rows.append(rec)

    with SESNSP_OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "cve_estado",
                "estado",
                "anio",
                "tasa_hom_doloso",
                "tasa_secuestro",
                "tasa_feminicidio",
                "tasa_extorsion",
                "fuente",
            ],
        )
        w.writeheader()
        w.writerows(rows)
    filled = sum(1 for r in rows if float(r["tasa_hom_doloso"]) > 0)
    print(f"SESNSP: {len(rows)} filas ({filled} con homicidio>0) → {SESNSP_OUT}")
    print("Nota: SESNSP víctimas cubre principalmente 2015–2025; 2010–2014 quedarán en 0.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--desde", type=int, default=2010)
    parser.add_argument("--hasta", type=int, default=2024)
    parser.add_argument("--solo-coneval", action="store_true")
    parser.add_argument("--solo-sesnsp", action="store_true")
    args = parser.parse_args()

    if not args.solo_sesnsp:
        process_coneval(args.desde, args.hasta)
    if not args.solo_coneval:
        process_sesnsp(args.desde, args.hasta)

    print("\nSiguiente paso: python3 build_state_panel.py")


if __name__ == "__main__":
    main()
